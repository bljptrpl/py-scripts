import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\dvillalva\\Desktop\\scripts_from_x\\architecture_doc.xlsx')

ws = wb.active
print("\n")
print(f'{ws["B3"].value} . {ws["C3"].value}     ->      ods . {ws["O3"].value} . {ws["P3"].value}')
print("\n")